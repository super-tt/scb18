"""
接口自动化测试:
1、excel测试用例准备ok，代码可以自动读取用例数据
2、执行接口测试，得到响应结果
3、断言:响应结果==预期结果     ---通过/不通过
4、写到最终执行通过与否的结果   --excel表格

"""
import requests
import openpyxl

# 1、excel测试用例准备ok，代码可以自动读取用例数据
def read_case(file_name, sheet_name):
    wb = openpyxl.load_workbook(file_name)  # 加载工作簿，打开一个excel文件
    sheet = wb[sheet_name]  # 打开某一个表单
    max_row = sheet.max_row  # 获取最大行数
    data_list = []
    # for i in range(5,15):
    for i in range(2, max_row + 1):
        dict1 = dict(
            case_id=sheet.cell(row=i, column=1).value,  # 获取case_id的值
            url=sheet.cell(row=i, column=5).value,  # 读取URL值
            data=sheet.cell(row=i, column=6).value,  # 读取data值
            expect=sheet.cell(row=i, column=7).value  # 读取期望
        )
        data_list.append(dict1)  # 把每一行读取到的测试用例数据生成的字典，追加到list中
    return data_list


# 2、执行接口测试，得到响应结果
def api_fun(url_login, data_login):
    headers_login = {'X-Lemonban-Media-Type': 'lemonban.v2', 'Content-Type': 'application/json'}  # 请求头
    response_login = requests.post(url=url_login, json=data_login, headers=headers_login).json()
    return response_login

# 4、写入结果到excel，定义成一个函数
def write_result(file_name, sheet_name, row, column, final_value):
    wb = openpyxl.load_workbook(file_name)  # 加载工作簿，打开一个excel文件
    sheet = wb[sheet_name]  # 打开名为login的sheet页
    sheet.cell(row=row, column=column).value = final_value  # 第二行第八列的值设置为passed
    wb.save(file_name)  # 保存文件（该文件不能为打开状态）

# 断言：实际结果==预期结果
# cases = read_case('test_case_api.xlsx','register')      # 调用函数，设置变量接收返回值
# for case in cases:
#     case_id = case['case_id']       # 获取第几条用例
#     url = case['url']       # 字符串格式
#     # data = case['data']         # 字符串格式，非字典
#     # expect = case['expect']         # 字符串格式，非字典
#     data = eval(case['data'])         # 字符串格式，转字典
#     expect = eval(case['expect'])         # 字符串格式，转字典
#     real_res = api_fun(url_login=url,data_login=data)
#     if (expect['code'] == real_res['code']) and (expect['msg'] == real_res['msg']):
#         # print("用例通过")
#         write_result('test_case_api.xlsx','register',case_id + 1, 8,'passed')
#     else:
#         # print("用例不通过")
#         write_result('test_case_api.xlsx', 'register', case_id + 1, 8, 'failed')




"""
eval() --------运行被字符串包裹着的表达式
'{"mobile_phone":"13652440101","pwd":"12345678","type":1,"reg_name":"34254sdfs"}' -----字符串里的字典
"""
# dict1 = eval('{"mobile_phone":"13652440101","pwd":"12345678","type":1,"reg_name":"34254sdfs"}')
# dict2 = eval('5+6')
# print(dict1)        # 字典格式：{'mobile_phone': '13652440101', 'pwd': '12345678', 'type': 1, 'reg_name': '34254sdfs'}
# print(dict2)        # 11


